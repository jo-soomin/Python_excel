import pandas as pd
from openpyxl import load_workbook
import openpyxl
import os
import xlsxwriter

source_path = 'C:/Users/user/PycharmProjects/IXYS/'

def pibot(ls_cmp_nm,ls_dt):
    df = pd.read_excel(source_path + 'result/'+ls_cmp_nm + ls_dt[0:8] +'_detail.xlsx', sheet_name='Summary')

    # 그 다음 피벗 테이블을 만드는 데 index(행) = YEAR, column(열) = Make 로 한다.
    #pivot = df.pivot_table(values="LOSS Q'ty", index= ('COMPANY', 'FAMILY', 'OPTION', 'PKG', 'PART NO', 'SALE CODE', 'RUN NO', 'LOT NO', 'MACHINE') , columns='LOSS NAME', aggfunc=np.sum, fill_value='')
    #print(pivot)

    # aggfunc를 두 개 이상 적용할 수 있다.
    # margin의 의미는 모든 값을 합산한 값을 보여주는 인덱스를 생성할 지를 결정하는 인자다.
    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(values="QTY", index= ('REL_NO', 'SALE_CD', 'DIE_CD', 'FAMILY', 'PKG'), columns='PRA_CD', aggfunc='sum', fill_value='', margins=True, margins_name='TOTAL')

    #pivot2_merge = df2 + df2_2


    '''
        df2_2 와 pivot2 테이블의 값중 PKG, SALE CODE 컬럼을 기준으로 공통된 값이 있으면 조합하여 병합한다. 
    '''
    pivot.to_excel(source_path + 'result/'+ls_cmp_nm + ls_dt[0:8] + '_detail_Pivot.xlsx', 'Pivot', header=True)
    '''
        excel 파일로 만듬
    '''
    filename = source_path + 'result/'+ls_cmp_nm + ls_dt[0:8] + '_detail_Pivot.xlsx'
    wb = load_workbook(filename)
    ws = wb.active

    ws.delete_rows(ws.max_row)

    wb.save(filename)


def pibot_unmerge(ls_cmp_nm, ls_dt):
    wb = load_workbook(source_path + 'result/'+ls_cmp_nm + ls_dt[0:8] + '_detail_Pivot.xlsx')
    '''
        엑셀 파일 열어서 wb변수에 저장
    '''
    sheets = wb.sheetnames  ##['Sheet1', 'Sheet2']

    for i,sheet in enumerate(sheets):
        '''enumerate : 인덱스 값 나열해줌'''
        ws = wb[sheets[i]]

        # you need a separate list to iterate on (see explanation #2 below)
        mergedcells =[]
        for group in ws.merged_cells.ranges:
            mergedcells.append(group)

        for group in mergedcells:
            min_col, min_row, max_col, max_row = group.bounds
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value

            # you need to unmerge before writing (see explanation #1 below)
            ws.unmerge_cells(str(group))

            for irow in range(min_row, max_row+1):
                for jcol in range(min_col, max_col+1):
                    ws.cell(row = irow, column = jcol, value = top_left_cell_value)

    wb.save(filename = source_path + 'result/'+ls_cmp_nm + ls_dt[0:8] + '_detail_Pivot.xlsx')




def item_pibot(ls_cmp_nm,ls_dt):
    df = pd.read_excel(source_path + 'IXYS20200918_item.xlsx', sheet_name='Summary')

    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(index='REL NO',
                           values=('OUT Q\'ty', 'VP', 'VFVR', 'Contact', 'IN Q\'ty', 'Isol.', 'VF', 'IR', 'VR'),
                           aggfunc='sum', fill_value='')

    # pivot2_merge = df2 + df2_2

    '''
        df2_2 와 pivot2 테이블의 값중 PKG, SALE CODE 컬럼을 기준으로 공통된 값이 있으면 조합하여 병합한다. 
    '''
    pivot.to_excel(source_path + 'IXYS20200918_item_pivot.xlsx', 'Pivot', header=True)

def rpt_pivot(ls_cmp_nm,ls_dt):
    df = pd.read_excel(source_path + 'result/IXYS20200918_rpt_l.xlsx', sheet_name='Summary')

    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(values='시료수', index=('구분','RUN'), columns='AA', aggfunc='sum', fill_value='')

    # pivot2_merge = df2 + df2_2

    '''
        df2_2 와 pivot2 테이블의 값중 PKG, SALE CODE 컬럼을 기준으로 공통된 값이 있으면 조합하여 병합한다. 
    '''
    pivot.to_excel(source_path + 'result/IXYS20200918_rpt_l_pivot.xlsx', 'Pivot', header=True)

def chart_pivot(ls_cmp_nm, ls_dt):
    df = pd.read_excel(source_path + 'IXYS20200918_total.xlsx', sheet_name='chart')

    pivot = df.pivot_table(values=('Package','Test in(units)','BIN1','VF','IR','VR','Contact','VP'), index='Package', aggfunc='sum', fill_value='')

    pivot.to_excel(source_path + 'result/IXYS20200918_total_pivot.xlsx', 'Pivot', header=True)

# 행 : Package, 값 : Test in(units), BIN1, VF, IR, VR, Contact, VP

def test_pibot(WW):
    df = df = pd.read_excel(source_path + 'result/IXYS(G)WW'+WW+'+test.xlsx', sheet_name='Summary')

    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(values="LOSS Q'ty", index=('PRA_CD','COMPANY', 'FAMILY', 'PKG','PKG OPT', 'PART NO', 'SALE CODE','DIE CODE','REL NO', 'RUN NO','SPLIT RUN NO', 'LOT NO', 'MAT YN', 'MCN CD' ,"IN Q'ty", "OUT Q'ty",'MV START DT','MV END DT','YIELD'), columns='LOSS NAME', aggfunc='sum')

    pivot.to_excel(os.path.abspath(source_path + 'result/IXYS(G)WW'+WW+'+test_pivot.xlsx'), 'Summary')

def test_pibot_unmerge(WW):
    wb = load_workbook(os.path.abspath(source_path + 'result/IXYS(G)WW'+WW+'+test_pivot.xlsx'))
    ws= wb.active
    '''
        엑셀 파일 열어서 wb변수에 저장
    '''
    sheets = wb.sheetnames  ##['Sheet1', 'Sheet2']

    for i,sheet in enumerate(sheets):
        '''enumerate : 인덱스 값 나열해줌'''
        ws = wb[sheets[i]]

        # you need a separate list to iterate on (see explanation #2 below)
        mergedcells =[]
        for group in ws.merged_cells.ranges:
            mergedcells.append(group)

        for group in mergedcells:
            min_col, min_row, max_col, max_row = group.bounds
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value

            # you need to unmerge before writing (see explanation #1 below)
            ws.unmerge_cells(str(group))

            for irow in range(min_row, max_row+1):
                for jcol in range(min_col, max_col+1):
                    ws.cell(row = irow, column = jcol, value = top_left_cell_value)
    ws.delete_cols(20)
    wb.save(os.path.abspath(source_path +'result/IXYS(G)WW'+WW+'+test_pivot_unmerge.xlsx'))


# IXYS(미국)-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def item_query_P_xlsx_pivot(A_source_path,CMP_NM,WW):
    df = df = pd.read_excel(A_source_path+CMP_NM+WW+'주차'+'.xlsx', sheet_name='Summary')

    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(values="LOSS Q'ty", index=('SALE CODE', 'SPLIT RUN NO', 'PKG', 'LOT NO', "IN Q'ty", "OUT Q'ty", 'MV END DT'), columns='LOSS NAME', aggfunc='sum')

    pivot.to_excel(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot'+'.xlsx'), 'Summary')

def item_query_P_xlsx_pivot_unmerge(A_source_path,CMP_NM,WW):
    wb = load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot'+'.xlsx'))
    ws= wb['Summary']
    '''
        엑셀 파일 열어서 wb변수에 저장
    '''
    sheets = wb.sheetnames  ##['Sheet1', 'Sheet2']

    for i,sheet in enumerate(sheets):
        '''enumerate : 인덱스 값 나열해줌'''
        ws = wb[sheets[i]]

        # you need a separate list to iterate on (see explanation #2 below)
        mergedcells =[]
        for group in ws.merged_cells.ranges:
            mergedcells.append(group)

        for group in mergedcells:
            min_col, min_row, max_col, max_row = group.bounds
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value

            # you need to unmerge before writing (see explanation #1 below)
            ws.unmerge_cells(str(group))

            for irow in range(min_row, max_row+1):
                for jcol in range(min_col, max_col+1):
                    ws.cell(row = irow, column = jcol, value = top_left_cell_value)

    for i in range(1, ws.max_column+1):
        if(ws.cell(row=1, column=i).value == ' '):
            ws.delete_cols(i)

    wb.save(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge'+'.xlsx'))
    wb.save(os.path.abspath(A_source_path + CMP_NM + WW + '주차_pivot_unmerge_data' + '.xlsx'))

def item_query_P_xlsx_pivot_unmerge_part_sum(A_source_path,CMP_NM,WW):
    df = df = pd.read_excel(A_source_path+CMP_NM+WW+'주차_pivot_unmerge'+'.xlsx', sheet_name='Summary')

    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(values=("IN Q'ty", "OUT Q'ty",'MV END DT','DC','UIS','Thermal'), index='aa', aggfunc='sum')

    pivot.to_excel(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge_part_sum'+'.xlsx'), 'Summary')

def item_query_T_xlsx_pivot(A_source_path,CMP_NM,WW):
    df = df = pd.read_excel(A_source_path+CMP_NM+WW+'주차 Test'+'.xlsx', sheet_name='Summary')

    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(index=('제품군', 'pkg', 'sale code', "split run no", 'lot no'), columns='pra_cd', values=('in qty','out qty'), aggfunc='sum')

    pivot.to_excel(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot'+'.xlsx'), 'Summary')

def item_query_T_xlsx_pivot_unmerge(A_source_path,CMP_NM,WW):
    wb = load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot'+'.xlsx'))
    ws= wb['Summary']
    '''
        엑셀 파일 열어서 wb변수에 저장
    '''
    sheets = wb.sheetnames  ##['Sheet1', 'Sheet2']

    for i,sheet in enumerate(sheets):
        '''enumerate : 인덱스 값 나열해줌'''
        ws = wb[sheets[i]]

        # you need a separate list to iterate on (see explanation #2 below)
        mergedcells =[]
        for group in ws.merged_cells.ranges:
            mergedcells.append(group)

        for group in mergedcells:
            min_col, min_row, max_col, max_row = group.bounds
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value

            # you need to unmerge before writing (see explanation #1 below)
            ws.unmerge_cells(str(group))

            for irow in range(min_row, max_row+1):
                for jcol in range(min_col, max_col+1):
                    ws.cell(row = irow, column = jcol, value = top_left_cell_value)

    for i in range(1, ws.max_column+1):
        if(ws.cell(row=1, column=i).value == ' '):
            ws.delete_cols(i)

    wb.save(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge'+'.xlsx'))


def item_query_T_xlsx_pivot_unmerge_sort(A_source_path,CMP_NM,WW):
    item_query_T_xlsx_pivot_unmerge = openpyxl.load_workbook(os.path.abspath(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge' + '.xlsx'))
    item_query_T_xlsx_pivot_unmerge_sheet = item_query_T_xlsx_pivot_unmerge.active
    item_query_T_xlsx_pivot_unmerge_sheet.delete_rows(1)
    all_values = []
    for row in item_query_T_xlsx_pivot_unmerge_sheet.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)

    item_query_T_xlsx_pivot_unmerge_res = xlsxwriter.Workbook(os.path.abspath(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge_res' + '.xlsx'))
    item_query_T_xlsx_pivot_unmerge_res_sheet = item_query_T_xlsx_pivot_unmerge_res.add_worksheet('Summary')

    for i in range(0, len(all_values)):
        for j in range(0, 19):
            item_query_T_xlsx_pivot_unmerge_res_sheet.write(i+1,j,sorted(all_values, key=lambda rows: rows[2])[i][j])
    item_query_T_xlsx_pivot_unmerge_res.close()

    item_query_T_xlsx_pivot_unmerge_res = openpyxl.load_workbook(os.path.abspath(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge_res' + '.xlsx'))
    item_query_T_xlsx_pivot_unmerge_res_sheet = item_query_T_xlsx_pivot_unmerge_res.active
    for i in range(1, 10+1):
        item_query_T_xlsx_pivot_unmerge_res_sheet.delete_cols(3)

    item_query_T_xlsx_pivot_unmerge_res.save(os.path.abspath(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge_res' + '.xlsx'))

def item_query_T_xlsx_pivot_unmerge_res_sort(A_source_path,CMP_NM,WW):
    item_query_T_xlsx_pivot_unmerge = openpyxl.load_workbook(os.path.abspath(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge_res' + '.xlsx'))
    item_query_T_xlsx_pivot_unmerge_sheet = item_query_T_xlsx_pivot_unmerge.active
    item_query_T_xlsx_pivot_unmerge_sheet.delete_rows(1)
    all_values = []
    for row in item_query_T_xlsx_pivot_unmerge_sheet.rows:
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)

    item_query_T_xlsx_pivot_unmerge_res = xlsxwriter.Workbook(os.path.abspath(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge_res_sort' + '.xlsx'))
    item_query_T_xlsx_pivot_unmerge_res_sheet = item_query_T_xlsx_pivot_unmerge_res.add_worksheet('Summary')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 0, 'SALE CD')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 1,'SPLIT RUN NO')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 2, 'ASSY IN')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 3, 'ASSY OUT')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 4, 'SOLDER')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 5, 'W/B')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 6, 'MOLD')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 7, 'TRIM')
    item_query_T_xlsx_pivot_unmerge_res_sheet.write(0, 8, 'OTHER')

    for i in range(0, len(all_values)):
        for j in range(0, 9):
            item_query_T_xlsx_pivot_unmerge_res_sheet.write(i + 1, j,sorted(all_values, key=lambda rows: (rows[0],rows[1]))[i][j])
    item_query_T_xlsx_pivot_unmerge_res.close()

def item_query_P_xlsx_pivot_unmerge_res_sort_part_sum(A_source_path,CMP_NM,WW):
    df = df = pd.read_excel(A_source_path + CMP_NM + WW + '주차_test_pivot_unmerge_res_sort' + '.xlsx', sheet_name='Summary')
    '''
        pivot_table 만들기
        index = 선언된 컬럼 기준으로 group by 한다. (행)
        aggfunc = group by 된 값들을 어떤 수식으로 표현할 것인지 설정
        values = 선언된 컬럼의 값만 출력 
        columns = 선언된 컬럼의 값을 열로 지정 (열)
    '''
    pivot = df.pivot_table(values=('ASSY IN', 'ASSY OUT', 'SOLDER', 'W/B', 'MOLD', 'TRIM', 'OTHER'), index='SPLIT RUN NO', aggfunc='sum')

    pivot.to_excel(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge_res_sort_part_sum' + '.xlsx'), 'Summary')
