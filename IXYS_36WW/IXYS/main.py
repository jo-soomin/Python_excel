import os
import xlsxwriter
import openpyxl
from db import sysdate
import db
from excel import pibot, item_pibot, rpt_pivot, chart_pivot, test_pibot, test_pibot_unmerge, item_query_P_xlsx_pivot, item_query_P_xlsx_pivot_unmerge, item_query_P_xlsx_pivot_unmerge_part_sum, item_query_T_xlsx_pivot, item_query_T_xlsx_pivot_unmerge, item_query_T_xlsx_pivot_unmerge_sort,item_query_T_xlsx_pivot_unmerge_res_sort, item_query_P_xlsx_pivot_unmerge_res_sort_part_sum
import datetime
from openpyxl.styles import PatternFill, Border, Side


source_path = 'C:/Users/user/PycharmProjects/IXYS/'



for ls_before_dt, ls_dt in sysdate():
    date = sysdate()

'''
ls_cmp_cd = '002011'
ls_before_dt = '20200918080000' 
ls_dt = '20200918080000'
'''
END_dt=datetime.date.today()
ST_dt=END_dt-datetime.timedelta(days=7)

year = str(END_dt)[0:4]
month = str(END_dt)[5:7]
day = str(END_dt)[8:10]
WW = datetime.date(int(year), int(month), int(day)).strftime("%V")

ST_dt = "20200828" + "080000"
END_dt = "20200904"+"080000"

rpt_l_ST_dt = "20200828" + "000000"
rpt_l_END_dt = "20200904" + "240000"

ls_cmp_cd = '002011'

if ls_cmp_cd == '002011':
    ls_cmp_nm ='IXYS'


result = db.detail_row(ST_dt,END_dt)
result2 = db.item_query(ST_dt,END_dt)
result_rpt_l_row = db.rpt_l_row(rpt_l_ST_dt,rpt_l_END_dt)

testwb = xlsxwriter.Workbook(os.path.abspath(source_path+'result/IXYS(G)WW'+WW+'+test.xlsx'))
testws = testwb.add_worksheet('Summary')

testws.write(0, 0, 'PRA_CD')
testws.write(0, 1, 'COMPANY')
testws.write(0, 2, 'FAMILY')
testws.write(0, 3, "PKG")
testws.write(0, 4, "PKG OPT")
testws.write(0, 5, "PART NO")
testws.write(0, 6, "SALE CODE")
testws.write(0, 7, "DIE CODE")
testws.write(0, 8, "REL NO")
testws.write(0, 9, "RUN NO")
testws.write(0, 10, "SPLIT RUN NO")
testws.write(0, 11, "LOT NO")
testws.write(0, 12, "MAT YN")
testws.write(0, 13, "MCN CD")
testws.write(0, 14, "IN Q'ty")
testws.write(0, 15, "OUT Q'ty")
testws.write(0, 16, "MV START DT")
testws.write(0, 17, "MV END DT")
testws.write(0, 18, "YIELD")
testws.write(0, 19, "LOSS NAME")
testws.write(0, 20, "LOSS Q'ty")

num = 0

for a, b, c, d, e, f, g, h, i, j, k, l, m, n, o, p, q, r, s, t, u in result2:
    num = num + 1

    testws.set_row(num, 20)  # Set the height
    '''
    num 열의 높이를 20으로 지정
    '''
    testws.write(num, 0, a)
    '''
    write : num행 0열에 컬럼 a 작성, cell_format_data 서식 적용
    '''
    testws.write(num, 1, b)
    testws.write(num, 2, c)
    testws.write(num, 3, d)
    if(e == None):
        testws.write(num, 4, ' ')
    else:
        testws.write(num, 4, e)
    testws.write(num, 5, f)
    testws.write(num, 6, g)
    testws.write(num, 7, h)
    testws.write(num, 8, i)
    testws.write(num, 9, j)
    testws.write(num, 10, k)
    testws.write(num, 11, l)
    testws.write(num, 12, m)
    testws.write(num, 13, n)
    testws.write(num, 14, o)
    testws.write(num, 15, p)
    testws.write(num, 16, q)
    testws.write(num, 17, r)
    testws.write(num, 18, s)
    if (t == None):
        testws.write(num, 19, '-')
    else:
        testws.write(num, 19, t)
    testws.write(num, 20, u)

# wb.close()

testwb.close()

test_pibot(WW)
test_pibot_unmerge(WW)

wb = xlsxwriter.Workbook(os.path.abspath(source_path+'/result/' + ls_cmp_nm + ls_dt[0:8]+ '_detail.xlsx'))
ws = wb.add_worksheet('Summary')

wb3 = xlsxwriter.Workbook(os.path.abspath(source_path+'/result/' + ls_cmp_nm + ls_dt[0:8]+ '_rpt_l.xlsx'))
ws3 = wb3.add_worksheet('Summary')

wbSum = xlsxwriter.Workbook(os.path.abspath(source_path+'/result/' + ls_cmp_nm + ls_dt[0:8]+ '_total.xlsx'))
wsSum = wbSum.add_worksheet('yield')
wsSum2 = wbSum.add_worksheet('Summary')
WsSum3 = wbSum.add_worksheet('chart')

#Title Width Set
'''
0행 0열의 폭(행넓이)을 14로 한다
'''

ws.set_column(0, 0, 14)
ws.set_column(1, 1, 9)
ws.set_column(2, 2, 14)
ws.set_column(3, 3, 10)
ws.set_column(4, 4, 25)
ws.set_column(5, 5, 25)
ws.set_column(6, 6, 15)
ws.set_column(7, 7, 14)
ws.set_column(8, 8, 11)
ws.set_column(9, 9, 11)
ws.set_column(10, 10, 11)
ws.set_column(11, 11, 11)
ws.set_column(12, 12, 25)
ws.set_column(13, 13, 11)


'''
add_format : 서식 적용
'''
#Cell1 Format(title)
cell_format_title = wb.add_format({'bold': True, 'align': 'center',  'border': 1, 'text_wrap': True, 'font_size' : 10})
cell_format_title.set_align('center')
cell_format_title.set_align('vcenter')
#Cell1 Format(data)
cell_format_data = wb.add_format({'bold': False, 'align': 'center',  'border': 1, 'text_wrap': True, 'font_size' : 10})
cell_format_data.set_align('center')
cell_format_data.set_align('vcenter')

#Cell3 Format(title)
cell_format_title3 = wb3.add_format({'bold': True, 'align': 'center',  'border': 1, 'text_wrap': True, 'font_size' : 10})
cell_format_title3.set_align('center')
cell_format_title3.set_align('vcenter')
#Cell3 Format(data)
cell_format_data3 = wb3.add_format({'bold': False, 'align': 'center',  'color': 'black', 'border': 1, 'text_wrap': True, 'font_size' : 10})
cell_format_data3.set_align('center')
cell_format_data3.set_align('vcenter')

#CellTotal Format(title)
cell_format_title4 = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1})
cell_format_title4.set_align('left')
cell_format_title4.set_align('vcenter')
#CellTotal Format(data)
cell_format_data4 = wbSum.add_format({'bold': False, 'border': 1, 'text_wrap': True, 'font_size' : 10})
cell_format_data4.set_align('left')
cell_format_data4.set_align('vcenter')

#total_title Format(title)
wbSum_B_L_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ffff99'})
wbSum_B_L_title.set_align('left')
wbSum_B_L_title.set_align('vcenter')

#total_title Format(title)
wbSum_M_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ccffcc'})
wbSum_M_title.set_align('left')
wbSum_M_title.set_align('vcenter')

#total_title Format(title)
wbSum_N_S_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ffcc99'})
wbSum_N_S_title.set_align('left')
wbSum_N_S_title.set_align('vcenter')

#total_title Format(title)
wbSum_T_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#00ccff'})
wbSum_T_title.set_align('left')
wbSum_T_title.set_align('vcenter')

#total_title Format(data)
wbSum_T_data = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#00ccff'})
wbSum_T_data.set_align('left')
wbSum_T_data.set_align('vcenter')

#total_title Format(title)
wbSum_U_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ccffcc'})
wbSum_U_title.set_align('left')
wbSum_U_title.set_align('vcenter')

#total_title Format(title)
wbSum_V_Y_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#78ff78'})
wbSum_V_Y_title.set_align('left')
wbSum_V_Y_title.set_align('vcenter')

#total_title Format(title)
wbSum_Z_AH_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ffcc99'})
wbSum_Z_AH_title.set_align('left')
wbSum_Z_AH_title.set_align('vcenter')

#total_title Format(title)
wbSum_AI_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#00ccff'})
wbSum_AI_title.set_align('left')
wbSum_AI_title.set_align('vcenter')

#total_title Format(title)
wbSum_AJ_AK_title = wbSum.add_format({'bold': True,'font_size' : 8, 'border': 1, 'fg_color': '#ff9900'})
wbSum_AJ_AK_title.set_align('left')
wbSum_AJ_AK_title.set_align('vcenter')

#total_title Format(title)
wbSum_AL_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ccffcc'})
wbSum_AL_title.set_align('left')
wbSum_AL_title.set_align('vcenter')

#total_title Format(title)
wbSum_AM_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ffff99'})
wbSum_AM_title.set_align('left')
wbSum_AM_title.set_align('vcenter')

#total_title Format(title)
wbSum_AN_AO_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ccffff'})
wbSum_AN_AO_title.set_align('left')
wbSum_AN_AO_title.set_align('vcenter')

#total_title Format(title)
wbSum_AP_BD_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'fg_color': '#ffff99'})
wbSum_AP_BD_title.set_align('left')
wbSum_AP_BD_title.set_align('vcenter')


#Total Summary Format(title)
Test_Summary_title = wbSum.add_format({'bold': True,'font_size' : 10, 'border': 1, 'align': 'center', 'font_size' : 10, 'fg_color': '#ebf1de'})

#Total Summary Format(data)
cell_format_data5 = wbSum.add_format({'bold': True, 'border': 1, 'text_wrap': True, 'font_size' : 9})


# Summary Width Set
wsSum2.set_column(0,0,0.8)
wsSum2.set_column(1,1,13)
wsSum2.merge_range('B6:D6',"WW"+WW+" Test yield summary")
wsSum2.hide_gridlines(2)

# Total Summary 머리글
wsSum2.write(0, 1, 'Document No : SPT#IXYS(G) test yield-'+datetime.date(int(year), int(month), int(day)).strftime('%y')+month+day)
wsSum2.write(1, 1, 'Date :'+datetime.date(int(year), int(month), int(day)).strftime('%b')+'/'+day+'th, '+year)
wsSum2.write(2, 1, 'Written by : SH.Kim in SP semi & com')
wsSum2.write(3, 1, 'Title : IXYS Weekly test yield report_ww'+WW)

# Total Summary 내용 title
wsSum2.write(6, 1, 'Package', Test_Summary_title)
wsSum2.write(6, 2, 'IN', Test_Summary_title)
wsSum2.write(6, 3, 'OUT', Test_Summary_title)
wsSum2.write(6, 4, "Yield", Test_Summary_title)
wsSum2.write(6, 5, "VF", Test_Summary_title)
wsSum2.write(6, 6, "IR", Test_Summary_title)
wsSum2.write(6, 7, "VR", Test_Summary_title)
wsSum2.write(6, 8, "Contact", Test_Summary_title)
wsSum2.write(6, 9, "VP", Test_Summary_title)


#상단 Title
'''
write : 0행 0열에 'COMPANY' 작성, cell_format_title 서식 적용
'''
ws.write(0, 0, 'REL_NO', cell_format_title)
ws.write(0, 1, 'SALE_CD', cell_format_title)
ws.write(0, 2, "PKG", cell_format_title)
ws.write(0, 3, "DIE_CD", cell_format_title)
ws.write(0, 4, "FAMILY", cell_format_title)
ws.write(0, 5, "PRA_CD", cell_format_title)
ws.write(0, 6, "QTY", cell_format_title)

num = 0

for a, b, c, d, e, f, g in result:
    num = num + 1
    ws.set_row(num, 30)  # Set the height
    '''
    num 열의 높이를 20으로 지정
    '''
    ws.write(num, 0, a, cell_format_data)
    '''
    write : num행 0열에 컬럼 a 작성, cell_format_data 서식 적용
    '''
    ws.write(num, 1, b, cell_format_data)
    ws.write(num, 2, c, cell_format_data)
    ws.write(num, 3, d, cell_format_data)
    ws.write(num, 4, e, cell_format_data)
    ws.write(num, 5, f, cell_format_data)
    ws.write(num, 6, g, cell_format_data)

wb.close()

ws3.write(0, 0, '구분', cell_format_title3)
ws3.write(0, 1, '고객', cell_format_title3)
ws3.write(0, 2, "PKG", cell_format_title3)
ws3.write(0, 3, "OPTION", cell_format_title3)
ws3.write(0, 4, "DEVICE", cell_format_title3)
ws3.write(0, 5, "RUN NO", cell_format_title3)
ws3.write(0, 6, "RUN 수량", cell_format_title3)
ws3.write(0, 7, "LOT NO", cell_format_title3)
ws3.write(0, 8, "고객 LOT NO", cell_format_title3)
ws3.write(0, 9, "시료수", cell_format_title3)
ws3.write(0, 10, "불량수", cell_format_title3)
ws3.write(0, 11, "불량명", cell_format_title3)
ws3.write(0, 12, "부적합구분", cell_format_title3)
ws3.write(0, 13, "부적합번호", cell_format_title3)
ws3.write(0, 14, "MMCL 구분", cell_format_title3)
ws3.write(0, 15, "정식구분", cell_format_title3)
ws3.write(0, 16, "판정여부", cell_format_title3)
ws3.write(0, 17, "설비번호", cell_format_title3)
ws3.write(0, 18, "검사일", cell_format_title3)
ws3.write(0, 19, "사원명", cell_format_title3)
ws3.write(0, 20, "근무조", cell_format_title3)
ws3.write(0, 21, "주코드", cell_format_title3)
ws3.write(0, 22, "품목구분", cell_format_title3)
ws3.write(0, 23, "이력수정자", cell_format_title3)
ws3.write(0, 24, "이력수정일", cell_format_title3)
ws3.write(0, 25, "비고", cell_format_title3)

num = 0

for a, b, c, d, e, f, g, h, i, j, k, l, m, n, o, p, q, r, s, t, u, v, w, x, y, z in result_rpt_l_row:
    num = num + 1
    ws3.set_row(num, 30)  # Set the height
    '''
    num 열의 높이를 30으로 지정
    '''
    if a == '2':
        ws3.write(num, 0, '외관', cell_format_data3)
    else:
        ws3.write(num, 0, '특성', cell_format_data3)
    if b =='002011':
        ws3.write(num, 1, 'IXYS(독일)', cell_format_data3)
    else:
        ws3.write(num, 1, b, cell_format_data3)
    ws3.write(num, 2, c, cell_format_data3)
    ws3.write(num, 3, d, cell_format_data3)
    ws3.write(num, 4, e, cell_format_data3)
    ws3.write(num, 5, f, cell_format_data3)
    ws3.write(num, 6, g, cell_format_data3)
    ws3.write(num, 7, h, cell_format_data3)
    ws3.write(num, 8, i, cell_format_data3)
    ws3.write(num, 9, j, cell_format_data3)
    ws3.write(num, 10, k, cell_format_data3)
    ws3.write(num, 11, l, cell_format_data3)
    ws3.write(num, 12, m, cell_format_data3)
    ws3.write(num, 13, n, cell_format_data3)
    ws3.write(num, 14, o, cell_format_data3)
    ws3.write(num, 15, p, cell_format_data3)
    if q == 'Y':
        ws3.write(num, 16, '합격', cell_format_data3)
    else:
        ws3.write(num, 16, '불합격', cell_format_data3)
    ws3.write(num, 17, r, cell_format_data3)
    ws3.write(num, 18, s, cell_format_data3)
    ws3.write(num, 19, t, cell_format_data3)
    ws3.write(num, 20, u, cell_format_data3)
    ws3.write(num, 21, v, cell_format_data3)
    if w =='RB01':
        ws3.write(num, 22, '양산제품', cell_format_data3)
    else:
        ws3.write(num, 22, w, cell_format_data3)
    ws3.write(num, 23, x, cell_format_data3)
    ws3.write(num, 24, y, cell_format_data3)
    ws3.write(num, 25, z, cell_format_data3)
    '''
        write : num행 0열에 컬럼 a 작성, cell_format_data 서식 적용
    '''
wb3.close()

wsSum.set_column('A:A',2)
wsSum.set_row(0, 13)
wsSum.set_row(1, 25)
wsSum.set_row(2, 25)
wsSum.merge_range('B1:L2',"Yiled Report_2020("+WW+"WW)",wbSum_B_L_title)
wsSum.write(2, 1, 'Assembly\nsite', wbSum_B_L_title)
wsSum.set_column('C:C',4)
wsSum.write(2, 2, 'WW', wbSum_B_L_title)
wsSum.write(2, 3, 'Assy_FKT', wbSum_B_L_title)
wsSum.write(2, 4, 'Reception\ndate', wbSum_B_L_title)
wsSum.write(2, 5, 'Shipment\ndate', wbSum_B_L_title)
wsSum.write(2, 6, 'Product', wbSum_B_L_title)
wsSum.write(2, 7, 'Package', wbSum_B_L_title)
wsSum.write(2, 8, 'Die type', wbSum_B_L_title)
wsSum.write(2, 9, 'No. Dice', wbSum_B_L_title)
wsSum.write(2, 10, 'Input_TA\n(Die)', wbSum_B_L_title)
wsSum.write(2, 11, 'Balance\nDice', wbSum_B_L_title)
wsSum.merge_range('M1:M3',"Start qty\n(units)",wbSum_M_title)
wsSum.merge_range('N1:S1',"Assembly Rejects per operation",wbSum_N_S_title)
wsSum.merge_range('N2:N3',"Dicing",wbSum_N_S_title)
wsSum.set_column('N:N',5)
wsSum.merge_range('O2:O3',"Soldering",wbSum_N_S_title)
wsSum.merge_range('P2:P3',"Wire\nBonding",wbSum_N_S_title)
wsSum.merge_range('Q2:Q3',"Moulding",wbSum_N_S_title)
wsSum.merge_range('R2:R3',"Trimming\n&Forming",wbSum_N_S_title)
wsSum.merge_range('S2:S3',"Other\nPhysical\nrejects",wbSum_N_S_title)
wsSum.merge_range('T1:T3',"Assembly\nYield",wbSum_T_title)
wsSum.merge_range('U1:U3',"Test in\n(units)",wbSum_U_title)
wsSum.merge_range('V1:Y2',"Good units per BIN",wbSum_V_Y_title)
wsSum.write(2, 21, 'BIN1', wbSum_V_Y_title)
wsSum.set_column('W:W',4)
wsSum.set_column('X:X',4)
wsSum.set_column('Y:Y',4)
wsSum.write(2, 22, 'BIN2', wbSum_V_Y_title)
wsSum.write(2, 23, 'BIN3', wbSum_V_Y_title)
wsSum.write(2, 24, 'BIN4', wbSum_V_Y_title)
wsSum.merge_range('Z1:AH1',"Electric Rejects",wbSum_Z_AH_title)
wsSum.merge_range('Z2:Z3',"Gate",wbSum_Z_AH_title)
wsSum.merge_range('AA2:AA3',"VF",wbSum_Z_AH_title)
wsSum.merge_range('AB2:AB3',"IR",wbSum_Z_AH_title)
wsSum.merge_range('AC2:AC3',"VR",wbSum_Z_AH_title)
wsSum.merge_range('AD2:AD3',"Isoll.",wbSum_Z_AH_title)
wsSum.merge_range('AE2:AE3',"Contact",wbSum_Z_AH_title)
wsSum.merge_range('AF2:AF3',"VF\nVR",wbSum_Z_AH_title)
wsSum.write(1, 32, 'VP', wbSum_Z_AH_title)
wsSum.write(2, 32, 'BIN17', wbSum_Z_AH_title)
wsSum.merge_range('AH2:AH3',"Iso\nTest",wbSum_Z_AH_title)
wsSum.merge_range('AI1:AI3',"Electrical\nYield",wbSum_AI_title)
wsSum.write(0, 35, 'Others', wbSum_AJ_AK_title)
wsSum.write(0, 36, '', wbSum_AJ_AK_title)
wsSum.merge_range('AJ2:AJ3',"Lost devices\nnot reflected\nin rejected\nclasses",wbSum_AJ_AK_title)
wsSum.merge_range('AK2:AK3',"Good\nUnits\nnot\nshipped",wbSum_AJ_AK_title)
wsSum.merge_range('AL1:AL3',"Shipped\nQty",wbSum_AL_title)
wsSum.merge_range('AM1:AM3',"Shipped\nRejects",wbSum_AM_title)
wsSum.merge_range('AN1:AN3',"Global\nYield\nLine",wbSum_AN_AO_title)
wsSum.merge_range('AO1:AO3',"Global\nYield\nIXYS",wbSum_AN_AO_title)
wsSum.merge_range('AP1:AP3',"Low yield\nreport?",wbSum_AP_BD_title)
wsSum.merge_range('AQ1:AQ3',"Limits\nrelaxation\nby special\nrelease?",wbSum_AP_BD_title)
wsSum.merge_range('AR1:BC1',"QA_Results",wbSum_AP_BD_title)
wsSum.merge_range('AR2:AT2',"Visual",wbSum_AP_BD_title)
wsSum.merge_range('AU2:AW2',"Isolation",wbSum_AP_BD_title)
wsSum.merge_range('AX2:AZ2',"Electrical(T-25ºC)",wbSum_AP_BD_title)
wsSum.merge_range('BA2:BC2',"Electrical(Hot Test)",wbSum_AP_BD_title)
wsSum.write(2, 43, 'Sample\nSize', wbSum_AP_BD_title)
wsSum.write(2, 44, 'NºRejects', wbSum_AP_BD_title)
wsSum.write(2, 45, 'Action', wbSum_AP_BD_title)
wsSum.write(2, 46, 'Sample\nSize', wbSum_AP_BD_title)
wsSum.write(2, 47, 'NºRejects', wbSum_AP_BD_title)
wsSum.write(2, 48, 'Action', wbSum_AP_BD_title)
wsSum.write(2, 49, 'Sample\nSize', wbSum_AP_BD_title)
wsSum.write(2, 50, 'NºRejects', wbSum_AP_BD_title)
wsSum.write(2, 51, 'Action', wbSum_AP_BD_title)
wsSum.write(2, 52, 'Sample\nSize', wbSum_AP_BD_title)
wsSum.write(2, 53, 'NºRejects', wbSum_AP_BD_title)
wsSum.write(2, 54, 'Action', wbSum_AP_BD_title)
wsSum.merge_range('BD1:BD3',"Comments",wbSum_AP_BD_title)


# 종합 ws 하나에 열 지정해서 다 넣기
wbSum.close()
pibot(ls_cmp_nm, ls_dt)

wbt = openpyxl.load_workbook(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8]+ '_total.xlsx'))
test = wbt.active
total_Summary =wbt['Summary']
total_chart = wbt['chart']

wbt2 = openpyxl.load_workbook(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8]+ '_detail_Pivot.xlsx'))
test2 = wbt2['Pivot']
max_length = str(test2.max_row)
max_col_2 = test2.max_column


wbt3 = openpyxl.load_workbook(os.path.abspath(source_path + 'result/IXYS(G)WW'+WW+'+test_pivot_unmerge.xlsx'))
test3 = wbt3['Summary']
max_row = test3.max_row

for i in range(1,9):
    test3.delete_cols(1)

for i in range(1,6):
    test3.delete_cols(2)

for i in range(1,4):
    test3.delete_cols(4)

max_col = test3.max_column

num = 4
for i in range(1, test2.max_column+1):
    for j in range(1, test2.max_row):
        test['B' + str(j+3)] = 'SP'
        test['C' + str(j+3)] = WW
        test['N' + str(num)] = 0
        if(test2.cell(row=1, column=i).value =='REL_NO'):
            test['D' + str(j+3)] = test2.cell(row=j+1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'SALE_CD'):
            test['G' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'DIE_CD'):
            test['J' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'FAMILY'):
            test['I' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'PKG'):
            test['H' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == '01_Reception date'):
            test['E' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == '02_Shipment date'):
            test['F' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == '03_INPUT_QTY'):
            test['K' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
            test['M' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif(test2.cell(row=1, column=i).value =='A020'):
            test['O' + str(j+3)] = test2.cell(row=j+1, column=i).value
        elif(test2.cell(row=1, column=i).value =='A030'):
            test['P' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'A040'):
            test['Q' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'A060'):
            test['R' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'A050'):
            test['S' + str(j + 3)] = test2.cell(row=j + 1, column=i).value
        elif (test2.cell(row=1, column=i).value == 'Z_Shipped Qty'):
            test['AL' + str(j + 3)] = test2.cell(row=j + 1, column=i).value

for i in range(1, max_col+1):
    print(test3.cell(row=1, column=i).value)
    if(test3.cell(row=1, column=i).value == 'Forward Voltage High'):
        for j in range(2, max_row + 1):
            if (test3.cell(row=j, column=i).value is not None):
                VF = 0
                VF = VF + test3.cell(row=j, column=i).value
                test3.cell(row=j, column=max_col + 1).value = VF
            # else:
            #     test3.cell(row=j, column=max_col + 1).value = 0
    elif (test3.cell(row=1, column=i).value == 'Collector to Base Leakage Over' or test3.cell(row=1,column=i).value == 'Collector to Emitter Leakage Over' or test3.cell(row=1, column=i).value == 'Drain to Source Leakage' or test3.cell(row=1,column=i).value == 'Drain-Source Leakage'):
        for j in range(2, max_row + 1):
            if (test3.cell(row=j, column=i).value is not None):
                IR = 0
                if(test3.cell(row=j, column=max_col + 2).value is not None):
                    IR = IR + test3.cell(row=j, column=i).value + test3.cell(row=j, column=max_col + 2).value
                    test3.cell(row=j, column=max_col + 2).value = IR
                else:
                    IR = IR + test3.cell(row=j, column=i).value
                    test3.cell(row=j, column=max_col + 2).value = IR
    elif (test3.cell(row=1, column=i).value == 'Collector to Base Voltage down' or test3.cell(row=1, column=i).value == 'Collector to Emitter Voltage down' or test3.cell(row=1, column=i).value == 'Drain to Source Voltage down'):
        for j in range(2, max_row + 1):
            if (test3.cell(row=j, column=i).value is not None):
                VR = 0
                if (test3.cell(row=j, column=max_col + 3).value is not None):
                    VR = VR + test3.cell(row=j, column=i).value + test3.cell(row=j, column=max_col + 3).value
                    test3.cell(row=j, column=max_col + 3).value = VR
                else:
                    VR = VR + test3.cell(row=j, column=i).value
                    test3.cell(row=j, column=max_col + 3).value = VR
    elif (test3.cell(row=1, column=i).value == 'Contact Fail' or test3.cell(row=1, column=i).value == 'DC Fail' or test3.cell(row=1, column=i).value == 'Drop Unit' or test3.cell(row=1, column=i).value == 'Drop pkg' or test3.cell(row=1, column=i).value == 'ENG`R Sample' or test3.cell(row=1, column=i).value == 'Eng\'r Sample' or test3.cell(row=1, column=i).value == 'External' or test3.cell(row=1, column=i).value == 'Misplaced Mark' or test3.cell(row=1, column=i).value == 'Prior process defect' or test3.cell(row=1, column=i).value == 'Static drain to source on state Resistance'):
        for j in range(2, max_row + 1):
            if (test3.cell(row=j, column=i).value is not None):
                Contact = 0
                if (test3.cell(row=j, column=max_col + 5).value is not None):
                    Contact = Contact + test3.cell(row=j, column=i).value + test3.cell(row=j, column=max_col + 5).value
                    test3.cell(row=j, column=max_col + 5).value = Contact
                else:
                    Contact = Contact + test3.cell(row=j, column=i).value
                    test3.cell(row=j, column=max_col + 5).value = Contact
    elif (test3.cell(row=1, column=i).value == 'Chip Open Fail' or test3.cell(row=1, column=i).value == 'Chip Short Fail' or test3.cell(row=1, column=i).value == 'VP FAIL'):
        for j in range(2, max_row + 1):
            if (test3.cell(row=j, column=i).value is not None):
                VpFail = 0
                if (test3.cell(row=j, column=max_col + 7).value is not None):
                    VpFail = VpFail + test3.cell(row=j, column=i).value + test3.cell(row=j, column=max_col + 7).value
                    test3.cell(row=j, column=max_col + 7).value = VpFail
                else:
                    VpFail = VpFail + test3.cell(row=j, column=i).value
                    test3.cell(row=j, column=max_col + 7).value = VpFail

test3.cell(row=1, column=max_col+1).value = 'VF'
test3.cell(row=1, column=max_col+2).value = 'IR'
test3.cell(row=1, column=max_col+3).value = 'VR'
test3.cell(row=1, column=max_col+4).value = 'Isol.'
test3.cell(row=1, column=max_col+5).value = 'Contact'
test3.cell(row=1, column=max_col+6).value = 'VFVR'
test3.cell(row=1, column=max_col+7).value = 'VP'

for isol in range(2, max_row +1):
    test3.cell(row=isol, column=max_col + 4).value = 0
for VFVR in range(2, max_row +1):
    test3.cell(row=VFVR, column=max_col + 6).value = 0

for i in range(4,max_col+1):
    test3.delete_cols(4)

wbt3.save(os.path.abspath(source_path + ls_cmp_nm + ls_dt[0:8]+ '_item.xlsx'))

item_pibot(ls_cmp_nm,ls_dt)

wbt4 = openpyxl.load_workbook(os.path.abspath(source_path + ls_cmp_nm + ls_dt[0:8]+'_item_pivot.xlsx'))
test4 = wbt4['Pivot']
max_row_4= test4.max_row
max_col_4 = test4.max_column

num2 =4
cells4 =test4['B2':'J'+str(max_row_4)]
for Contact, In_Qty, IR, Isol, Out_Qty, VF, VFVR, VP, VR in cells4:
    test['AE' + str(num2)] = Contact.value
    test['U' + str(num2)] = In_Qty.value
    test.cell(row=num2, column=21).fill = PatternFill(start_color='ccffcc', end_color='ccffcc', fill_type='solid')
    test.cell(row=num2, column=21).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    test['AB' + str(num2)] = IR.value
    test['AD' + str(num2)] = Isol.value
    test['V' + str(num2)] = Out_Qty.value
    test['AA' + str(num2)] = VF.value
    test['AF' + str(num2)] = VFVR.value
    test['AG' + str(num2)] = VP.value
    test['AC' + str(num2)] = VR.value

    num2 = num2 +1

wbt5 = openpyxl.load_workbook(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8]+ '_rpt_l.xlsx'))
test5 = wbt5['Summary']
rpt_max_row = test5.max_row
rpt_max_col = test5.max_column
test5.move_range("A1:Z"+str(rpt_max_row), cols=1)

test5['A' + str(1)] = 'AA'
for i in range(2, rpt_max_row+1):
    test5['A' + str(i)] = 1

for i in range(1, 4+1):
    test5.delete_cols(3)

for i in range(1, 3+1):
    test5.delete_cols(4)

for i in range(1, 16+1):
    test5.delete_cols(5)

test5.cell(row=1, column=3).value = 'RUN'
for i in range(2, rpt_max_row+1):
    test5.cell(row=i, column=3).value = test5.cell(row=i, column=3).value[0:6]


wbt5.save(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8] + '_rpt_l.xlsx'))
rpt_pivot(ls_cmp_nm,ls_dt)

wbt6 = openpyxl.load_workbook(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8]+ '_rpt_l_pivot.xlsx'))
test6 = wbt6['Pivot']
rpt_pivot_max_row = test6.max_row
rpt_pivot_max_col = test6.max_column

total_max_row = test.max_row
total_max_col = test.max_column
print('총 컬럼 갯수:'+ str(total_max_col))
print('총 로우 갯수:'+ str(total_max_row))

for i in range(2, rpt_pivot_max_row):
    if(test6.cell(row=i, column=1).value =='특성'):
        count = rpt_pivot_max_row - ((rpt_pivot_max_row - i) +1)
        count2 = (rpt_pivot_max_row - i) +1
        # print(test6.cell(row=i, column=2).value)
        test6.move_range('A'+str(i)+':'+'C'+str(rpt_pivot_max_row), rows=-count+1, cols=4)

wbt6.save(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8]+ '_rpt_l_pivot.xlsx'))

for i in range(4, total_max_row+1):
    test['AR' + str(i)] = test.cell(row=i, column=4).value

rpt_l_pivot_num = 4
rpt_l_pivot_cell =test6['B2':'C'+str(rpt_pivot_max_row)]
for RUN, NUM in rpt_l_pivot_cell:
    test['BE' + str(rpt_l_pivot_num)] = RUN.value
    test['BF' + str(rpt_l_pivot_num)] = NUM.value
    rpt_l_pivot_num = rpt_l_pivot_num +1

rpt_l_pivot_num2 = 4
rpt_l_pivot_cell2 =test6['F2':'G'+str(rpt_pivot_max_row)]
for Characteristic, NUM in rpt_l_pivot_cell2:
    test['BG' + str(rpt_l_pivot_num2)] = Characteristic.value
    test['BH' + str(rpt_l_pivot_num2)] = NUM.value
    rpt_l_pivot_num2 = rpt_l_pivot_num2 +1

total_max_row2 = test.max_row
total_max_col2 = test.max_column

for i in range(4, total_max_row2+1):
    if(test.cell(row=i, column=4).value is not None):
        test.cell(row=i, column=44).value = None
        for j in range(4, total_max_row2+1):
            if(test.cell(row=i, column=4).value == test.cell(row=j, column=57).value):
                test.cell(row=i, column=44).value = test.cell(row=j, column=58).value
                test.cell(row=i, column=45).value = 0

for i in range(4, total_max_row2+1):
    if(test.cell(row=i, column=4).value is not None):
        test.cell(row=i, column=50).value = None
        for j in range(4, total_max_row2+1):
            if(test.cell(row=i, column=4).value == test.cell(row=j, column=59).value):
                test.cell(row=i, column=50).value = test.cell(row=j, column=60).value
                test.cell(row=i, column=51).value = 0

for i in range(total_max_col2, total_max_col2-4, -1):
    test.delete_cols(i)

sum = 0
sum2 = 0
sum3 = 0
for i in range(4,total_max_row+1):

    for j in range(14, 19+1):
        if(test.cell(row=i, column=j).value):
            sum = sum + test.cell(row=i, column=j).value
    # print(sum)

    for x in range(22, 34+1):
        if(test.cell(row=i, column=x).value is not None):
            sum2 = sum2 + test.cell(row=i, column=x).value
    # print(sum2)
    sum3 = sum + sum2
    # print(sum3)
    final = test.cell(row=i, column=13).value - sum3
    test.cell(row=i, column=36).value = final
    test.cell(row=i, column=15).value = test.cell(row=i, column=15).value + test.cell(row=i, column=36).value
    test.cell(row=i, column=36).value = 0
    test.cell(row=i, column=36).fill = PatternFill(start_color='ff9900', end_color='ff9900', fill_type='solid')
    test.cell(row=i, column=36).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    # print(final)
    sum = 0
    sum2 = 0
wbt.save(os.path.abspath(source_path + ls_cmp_nm + ls_dt[0:8] + '_total.xlsx'))

for i in range(4, total_max_row+1):
    test['N' + str(i)].value = 0

for i in range(4, total_max_row+1):
    try:

        if (test['AL' + str(i)].value == None):
            test['AL' + str(i)].value = 0
        if (test['AK' + str(i)].value == None):
            test['AK' + str(i)].value = 0
        test['AN' + str(i)].value = str(round((test['AL' + str(i)].value + test['AK' + str(i)].value) / test['M' + str(i)].value * 100, 1)) + '%'
        test.cell(row=i, column=40).fill = PatternFill(start_color='ccffff', end_color='ccffff', fill_type='solid')
        test.cell(row=i, column=40).border = Border(left=Side(style="thin"), right=Side(style="thin"),top=Side(style="thin"), bottom=Side(style="thin"))

        test['AO' + str(i)].value = str(round(((test['AL' + str(i)].value) / test['M' + str(i)].value) * 100, 1)) + '%'
        test.cell(row=i, column=41).fill = PatternFill(start_color='ccffff', end_color='ccffff', fill_type='solid')
        test.cell(row=i, column=41).border = Border(left=Side(style="thin"), right=Side(style="thin"),top=Side(style="thin"), bottom=Side(style="thin"))

        if (test['AL' + str(i)].value == 0):
            test['AL' + str(i)].value = None
        if (test['AK' + str(i)].value == 0):
            test['AK' + str(i)].value = None

        AI_SUM = 0
        for j in range(22, 25 + 1):
            if (test.cell(row=i, column=j).value != None):
                AI_SUM = AI_SUM + test.cell(row=i, column=j).value
        test['AI' + str(i)].value = str(round((AI_SUM / test['U' + str(i)].value) * 100, 2)) + '%'
        test.cell(row=i, column=35).fill = PatternFill(start_color='00ccff', end_color='00ccff', fill_type='solid')
        test.cell(row=i, column=35).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        T_SUM = 0
        for j in range(14, 19 + 1):
            if (test.cell(row=i, column=j).value != None):
                T_SUM = T_SUM + test.cell(row=i, column=j).value
            print(test.cell(row=i, column=j).value)
        test['T' + str(i)].value = str(round(((test['M' + str(i)].value - T_SUM) / test['M' + str(i)].value) * 100, 2)) + '%'
        test.cell(row=i, column=20).fill = PatternFill(start_color='00ccff', end_color='00ccff',fill_type='solid')
        test.cell(row=i, column=20).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    except ZeroDivisionError:
            print("ZeroDivision")

total_chart['A' + str(1)] = 'Package'
total_chart['B' + str(1)] = 'Test in(units)'
total_chart['C' + str(1)] = 'BIN1'
total_chart['D' + str(1)] = 'VF'
total_chart['E' + str(1)] = 'IR'
total_chart['F' + str(1)] = 'VR'
total_chart['G' + str(1)] = 'Contact'
total_chart['H' + str(1)] = 'VP'

for chart_sheet in range(4, total_max_row+1):
     total_chart['A' + str(chart_sheet-2)] = test['H' + str(chart_sheet)].value
     total_chart['B' + str(chart_sheet-2)] = test['U' + str(chart_sheet)].value
     total_chart['C' + str(chart_sheet-2)] = test['V' + str(chart_sheet)].value
     total_chart['D' + str(chart_sheet-2)] = test['AA' + str(chart_sheet)].value
     total_chart['E' + str(chart_sheet-2)] = test['AB' + str(chart_sheet)].value
     total_chart['F' + str(chart_sheet-2)] = test['AC' + str(chart_sheet)].value
     total_chart['G' + str(chart_sheet-2)] = test['AE' + str(chart_sheet)].value
     total_chart['H' + str(chart_sheet-2)] = test['AG' + str(chart_sheet)].value

wbt6.save(os.path.abspath(source_path + 'result/' + ls_cmp_nm + ls_dt[0:8]+ '_rpt_l_pivot.xlsx'))
wbt.save(os.path.abspath(source_path + ls_cmp_nm + ls_dt[0:8] + '_total.xlsx'))

chart_pivot(ls_cmp_nm,ls_cmp_cd)

wbt7 = openpyxl.load_workbook(os.path.abspath(source_path + '/result/' + ls_cmp_nm + ls_dt[0:8]+ '_total_pivot.xlsx'))
wbt7_sheet = wbt7['Pivot']
wbt7_sheet_max_row = wbt7_sheet.max_row
wbt7_sheet_max_col = wbt7_sheet.max_column


wbt7_sheet.cell(row=1, column=wbt7_sheet_max_col+1).value = 'YIELD'

Pkg = 0
BIN1 = 0
Test_in = 0
YIELD = 0
VF = 0
IR = 0
VR = 0
Contact = 0
VP = 0

for i in range(1, wbt7_sheet_max_row+1):
    if(i<2):
        for j in range(1, wbt7_sheet_max_col+1+1):
            if(wbt7_sheet.cell(row=i, column=j).value == 'Package'):
                Pkg =j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'BIN1'):
                BIN1 = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'Test in(units)'):
                Test_in = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'YIELD'):
                YIELD = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'VF'):
                VF = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'IR'):
                IR = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'VR'):
                VR = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'Contact'):
                Contact = j
            elif(wbt7_sheet.cell(row=i, column=j).value == 'VP'):
                VP = j
    else:
        wbt7_sheet.cell(row=i, column=wbt7_sheet_max_col + 1).value = str(round((wbt7_sheet.cell(row=i, column=BIN1).value) / (wbt7_sheet.cell(row=i, column=Test_in).value)*100,2)) + '%'

for i in range(2, wbt7_sheet_max_row+1):
    total_Summary.cell(row=i + 6, column=2).value = (wbt7_sheet.cell(row=i, column=Pkg).value)
    total_Summary.cell(row=i + 6, column=2).fill = PatternFill(start_color='ebf1de', end_color='ebf1de', fill_type='solid')
    total_Summary.cell(row=i + 6, column=3).value = (wbt7_sheet.cell(row=i, column=Test_in).value)
    total_Summary.cell(row=i + 6, column=4).value = (wbt7_sheet.cell(row=i, column=BIN1).value)
    total_Summary.cell(row=i + 6, column=5).value = (wbt7_sheet.cell(row=i, column=YIELD).value)
    total_Summary.cell(row=i + 6, column=6).value = (wbt7_sheet.cell(row=i, column=VF).value)
    total_Summary.cell(row=i + 6, column=7).value = (wbt7_sheet.cell(row=i, column=IR).value)
    total_Summary.cell(row=i + 6, column=8).value = (wbt7_sheet.cell(row=i, column=VR).value)
    total_Summary.cell(row=i + 6, column=9).value = (wbt7_sheet.cell(row=i, column=Contact).value)
    total_Summary.cell(row=i + 6, column=10).value = (wbt7_sheet.cell(row=i, column=VP).value)
    for j in range(2, 11):
        total_Summary.cell(row=i + 6, column=j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        total_Summary.cell(row=i + 6, column=j).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

total_Summary.cell(row=total_Summary.max_row+1, column=2).value ='total'

wbt7.save(os.path.abspath(source_path + 'result/' + ls_cmp_nm + ls_dt[0:8]+ '_total_pivot.xlsx'))
wbt.save(os.path.abspath(source_path + ls_cmp_nm + ls_dt[0:8] + '_total.xlsx'))

Summary_sum = 0
for i in range(3, total_Summary.max_column+1):
    for j in range(8, total_Summary.max_row):
        if(i != 5):
            Summary_sum = Summary_sum + total_Summary.cell(row=j, column=i).value
    total_Summary.cell(row=total_Summary.max_row, column=i).value = Summary_sum
    Summary_sum = 0
wbt.save(os.path.abspath(source_path + ls_cmp_nm + ls_dt[0:8] + '_total.xlsx'))
total_Summary.cell(row=total_Summary.max_row, column=5).value = str(round(((total_Summary.cell(row=total_Summary.max_row, column=4).value)/(total_Summary.cell(row=total_Summary.max_row, column=3).value))*100,2)) + '%'

for j in range(2, 10+1):
    total_Summary.cell(row=total_Summary.max_row, column=j).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
    total_Summary.cell(row=total_Summary.max_row, column=j).fill = PatternFill(start_color='f2dcdb', end_color='f2dcdb',fill_type='solid')
    total_Summary.cell(row=total_Summary.max_row, column=j).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

wbt.remove(wbt['chart'])
wbt.save(os.path.abspath(source_path + ls_cmp_nm +'(G) test yield repot_WW'+WW+'_2020.xlsx'))


# IXYS(미국)-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


# IXYS(미국) 회사 코드 (CMP_cd): 002009
# 공정 코드 ( PRA_cd ): T010
# 프로젝트 경로
A_source_path = 'C:/Users/user/PycharmProjects/IXYS/IXYS(미국)/'

# 기준 시작날짜 종료날짜
END_dt=datetime.date.today()
ST_dt=END_dt-datetime.timedelta(days=7)

#  주차 구하기위해 년도 월 일 분할
year = str(END_dt)[0:4]
month = str(END_dt)[5:7]
day = str(END_dt)[8:10]

# 주차 구하기
WW = datetime.date(int(year), int(month), int(day)).strftime("%V")

# 기준 시작날짜 종료날짜
ST_dt = str(ST_dt)[0:4] + str(ST_dt)[5:7]+str(ST_dt)[8:10] + "080000"
END_dt = str(END_dt)[0:4] + str(END_dt)[5:7]+str(END_dt)[8:10] + "080000"

# 고객번호, 고객이름, 공정번호
CMP_cd = '002009'
CMP_NM = 'IXYS(미국)'
PRA_cd = 'T010'


# 공정기준 아이템 수율현황 쿼리
item_query_P_result = db.item_query_P(ST_dt, END_dt)

item_query_P_xlsx = xlsxwriter.Workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차'+'.xlsx'))
item_query_P_xlsx_sheet = item_query_P_xlsx.add_worksheet('Summary')

item_query_P_xlsx_sheet.write(0, 0, 'PKG')
item_query_P_xlsx_sheet.write(0, 1, 'SALE CODE')
item_query_P_xlsx_sheet.write(0, 2, 'SPLIT RUN NO')
item_query_P_xlsx_sheet.write(0, 3, "LOT NO")
item_query_P_xlsx_sheet.write(0, 4, "IN Q'ty")
item_query_P_xlsx_sheet.write(0, 5, "OUT Q'ty")
item_query_P_xlsx_sheet.write(0, 6, "MV END DT")
item_query_P_xlsx_sheet.write(0, 7, "LOSS NAME")
item_query_P_xlsx_sheet.write(0, 8, "LOSS Q'ty")

num = 0
for a, b, c, d, e, f, g, h, i in item_query_P_result:
    num = num + 1

    item_query_P_xlsx_sheet.write(num, 0, a)
    '''
    write : num행 0열에 컬럼 a 작성, cell_format_data 서식 적용
    '''
    item_query_P_xlsx_sheet.write(num, 1, b)
    item_query_P_xlsx_sheet.write(num, 2, c)
    item_query_P_xlsx_sheet.write(num, 3, d)
    item_query_P_xlsx_sheet.write(num, 4, e)
    item_query_P_xlsx_sheet.write(num, 5, f)
    item_query_P_xlsx_sheet.write(num, 6, g)
    if(h is None):
        item_query_P_xlsx_sheet.write(num, 7, ' ')
    else:
        item_query_P_xlsx_sheet.write(num, 7, h)
    item_query_P_xlsx_sheet.write(num, 8, i)

item_query_P_xlsx.close()

item_query_P_xlsx_pivot(A_source_path,CMP_NM,WW)
item_query_P_xlsx_pivot_unmerge(A_source_path,CMP_NM,WW)

IXYS_USA_pivot_unmerge = openpyxl.load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge'+'.xlsx'))
IXYS_USA_pivot_unmerge_sheet = IXYS_USA_pivot_unmerge.active
IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=IXYS_USA_pivot_unmerge_sheet.max_column+1, value='DC')
IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=IXYS_USA_pivot_unmerge_sheet.max_column+1, value='UIS')
IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=IXYS_USA_pivot_unmerge_sheet.max_column+1, value='Thermal')
IXYS_USA_pivot_unmerge.save(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge'+'.xlsx'))

for i in range(2, IXYS_USA_pivot_unmerge_sheet.max_row+1):
    DC = 0
    UIS = 0
    Thermal = 0
    for j in range(1, IXYS_USA_pivot_unmerge_sheet.max_column+1):
        if(IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Base to Collector Saturation' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Contact Fail' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'DC Fail' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Drop Unit' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Drop pkg' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'ENG`R Sample' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Eng\'r Sample' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'External' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Misplaced Mark' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Prior process defect' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'X Mark'):
            if(IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=j).value is not None):
                DC = DC + IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=j).value
                IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column - 2).value = DC
            elif(IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column - 2).value is None):
                IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column - 2).value = 0
        elif(IXYS_USA_pivot_unmerge_sheet.cell(row=1,column=j).value == 'Chip Open Fail' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Single Avalanche Voltage'):
            if (IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=j).value is not None):
                UIS = UIS + IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=j).value
                IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column - 1).value = UIS
            elif(IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column - 1).value is None):
                IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column - 1).value = 0
        elif (IXYS_USA_pivot_unmerge_sheet.cell(row=1,column=j).value == 'BVCBO(2)-BVCBO(1)' or IXYS_USA_pivot_unmerge_sheet.cell(row=1, column=j).value == 'Base to Emitter Delta Volt'):
            if (IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=j).value is not None):
                Thermal = Thermal + IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=j).value
                IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column).value = Thermal
            elif(IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column).value is None):
                IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=IXYS_USA_pivot_unmerge_sheet.max_column).value = 0

for i in range(1, ((IXYS_USA_pivot_unmerge_sheet.max_column-3)-7)+1):
    IXYS_USA_pivot_unmerge_sheet.delete_cols(8)

for i in range(1, IXYS_USA_pivot_unmerge_sheet.max_row+1):
    if(i == 1):
        IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=4).value = 'aa'
    else:
        IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=4).value = IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=1).value + "," +IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=2).value + "," + IXYS_USA_pivot_unmerge_sheet.cell(row=i, column=3).value

for i in range(1, 3+1):
    IXYS_USA_pivot_unmerge_sheet.delete_cols(1)

IXYS_USA_pivot_unmerge.save(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge'+'.xlsx'))

item_query_P_xlsx_pivot_unmerge_part_sum(A_source_path,CMP_NM,WW)

IXYS_USA_pivot_unmerge_part_sum = openpyxl.load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge_part_sum'+'.xlsx'))
IXYS_USA_pivot_unmerge_part_sum_sheet = IXYS_USA_pivot_unmerge_part_sum.active

IXYS_USA_pivot_unmerge_part_sum_sheet.move_range("B1:F"+str(IXYS_USA_pivot_unmerge_part_sum_sheet.max_row), cols=2)

for i in range(1, IXYS_USA_pivot_unmerge_part_sum_sheet.max_row+1):
    if(i == 1):
        IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=1).value = 'sale'
        IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=2).value = 'run'
        IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=3).value = 'pkg'
    else:
        str_sum = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=1).value
        IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=1).value = str_sum.split(",")[0]
        IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=2).value = str_sum.split(",")[1]
        IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=3).value = str_sum.split(",")[2]

IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=1, column=9).value = 'aa'
IXYS_USA_pivot_unmerge_data = openpyxl.load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge_data'+'.xlsx'))
IXYS_USA_pivot_unmerge_data_sheet = IXYS_USA_pivot_unmerge_data.active

for i in range(2, IXYS_USA_pivot_unmerge_part_sum_sheet.max_row+1):
    list = []
    for j in range(2, IXYS_USA_pivot_unmerge_data_sheet.max_row+1):
        if(IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=2).value == IXYS_USA_pivot_unmerge_data_sheet.cell(row=j, column=2).value):
            list.append(IXYS_USA_pivot_unmerge_data_sheet.cell(row=j, column=7).value)
            IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=9).value = list[0]

IXYS_USA_pivot_unmerge_part_sum.save(os.path.abspath(A_source_path+CMP_NM+WW+'주차_pivot_unmerge_part_sum'+'.xlsx'))

# IXYS 미국 total excel
IXYS_A_Total = xlsxwriter.Workbook(os.path.abspath(A_source_path+'IXYS test yield repot_WW'+WW+'_2020.xlsx'))
IXYS_A_Total_sheet = IXYS_A_Total.add_worksheet('MASS+ER')

# IXYS 미국 tottal title format
IXYS_A_Total_title = IXYS_A_Total.add_format({'align': 'center', 'text_wrap': True, 'font_size': 9, 'bg_color':'#ffd32a'})

IXYS_A_Total_sheet.write(0, 0, 'Date', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 1, 'Device', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 2, 'PKG', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 3, 'RunNo', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 4, 'Assy in', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 5, 'Assy out', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 6, 'Assy yield', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 7, 'Solder', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 8, 'W/B', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 9, 'Mold', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 10, 'Trim/Form', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 11, 'Other', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 12, 'Test In', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 13, 'Test Out', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 14, 'Test yield', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 15, 'DC', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 16, 'UIS', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 17, 'Thermal', IXYS_A_Total_title)
IXYS_A_Total_sheet.write(0, 18, 'Failure Mode', IXYS_A_Total_title)
IXYS_A_Total_sheet.set_column('A:A',9.8)
IXYS_A_Total_sheet.set_column('B:B',17)
IXYS_A_Total_sheet.set_column('C:C',14)
IXYS_A_Total_sheet.set_column('D:D',14)
IXYS_A_Total_sheet.set_column('S:S',10)

IXYS_A_Total.close()

IXYS_A_Total = openpyxl.load_workbook(os.path.abspath(A_source_path+'IXYS test yield repot_WW'+WW+'_2020.xlsx'))
IXYS_A_Total_sheet = IXYS_A_Total.active

for i in range(2, IXYS_USA_pivot_unmerge_part_sum_sheet.max_row+1):
    IXYS_A_Total_sheet.cell(row=i, column=1).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=9).value
    IXYS_A_Total_sheet.cell(row=i, column=2).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=1).value
    IXYS_A_Total_sheet.cell(row=i, column=3).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=3).value
    IXYS_A_Total_sheet.cell(row=i, column=4).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=2).value
    IXYS_A_Total_sheet.cell(row=i, column=13).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=5).value
    IXYS_A_Total_sheet.cell(row=i, column=14).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=6).value
    IXYS_A_Total_sheet.cell(row=i, column=16).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=4).value
    IXYS_A_Total_sheet.cell(row=i, column=17).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=8).value
    IXYS_A_Total_sheet.cell(row=i, column=18).value = IXYS_USA_pivot_unmerge_part_sum_sheet.cell(row=i, column=7).value

item_query_T_result = db.item_query_T(ST_dt, END_dt)

# Item별 수율현황 Test 기준 수율
item_query_T_xlsx = xlsxwriter.Workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차 Test'+'.xlsx'))
item_query_T_xlsx_sheet = item_query_T_xlsx.add_worksheet('Summary')

num =0
for a,b,c,d,e,f,g,h,i,j,k,l in item_query_T_result:
    num = num+1
    item_query_T_xlsx_sheet.write(num, 0, a)
    item_query_T_xlsx_sheet.write(num, 1, b)
    item_query_T_xlsx_sheet.write(num, 2, c)
    item_query_T_xlsx_sheet.write(num, 3, d)
    item_query_T_xlsx_sheet.write(num, 4, e)
    item_query_T_xlsx_sheet.write(num, 5, f)
    item_query_T_xlsx_sheet.write(num, 6, g)
    item_query_T_xlsx_sheet.write(num, 7, h)
    item_query_T_xlsx_sheet.write(num, 8, i)
    item_query_T_xlsx_sheet.write(num, 9, j)
    item_query_T_xlsx_sheet.write(num, 10, k)
    item_query_T_xlsx_sheet.write(num, 11, l)

item_query_T_xlsx_sheet.write(0, 0, 'pra_cd')
item_query_T_xlsx_sheet.write(0, 2, '제품군')
item_query_T_xlsx_sheet.write(0, 3, 'pkg')
item_query_T_xlsx_sheet.write(0, 6, 'sale code')
item_query_T_xlsx_sheet.write(0, 8, 'split run no')
item_query_T_xlsx_sheet.write(0, 9, 'lot no')
item_query_T_xlsx_sheet.write(0, 10, 'in qty')
item_query_T_xlsx_sheet.write(0, 11, 'out qty')

item_query_T_xlsx.close()

# Item별 수율현황 Test 기준 수율 pivot
item_query_T_xlsx_pivot(A_source_path,CMP_NM,WW)
# Item별 수율현황 Test 기준 수율 unmerge
item_query_T_xlsx_pivot_unmerge(A_source_path,CMP_NM,WW)

item_query_T_xlsx_pivot_unmerge = openpyxl.load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge'+'.xlsx'))
item_query_T_xlsx_pivot_unmerge_sheet = item_query_T_xlsx_pivot_unmerge.active

for i in range(1, item_query_T_xlsx_pivot_unmerge_sheet.max_column+1):
    if(item_query_T_xlsx_pivot_unmerge_sheet.cell(row=2, column=i).value == 'A020' or
       item_query_T_xlsx_pivot_unmerge_sheet.cell(row=2, column=i).value == 'A030' or
       item_query_T_xlsx_pivot_unmerge_sheet.cell(row=2, column=i).value == 'A040' or
       item_query_T_xlsx_pivot_unmerge_sheet.cell(row=2, column=i).value == 'A050' or
       item_query_T_xlsx_pivot_unmerge_sheet.cell(row=2, column=i).value == 'A060'):
        item_query_T_xlsx_pivot_unmerge_sheet.cell(row=3, column=i).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=2, column=i).value

item_query_T_xlsx_pivot_unmerge_sheet.delete_rows(1)
item_query_T_xlsx_pivot_unmerge_sheet.delete_rows(1)
item_query_T_xlsx_pivot_unmerge_sheet.delete_cols(1)
item_query_T_xlsx_pivot_unmerge_sheet.delete_cols(1)
item_query_T_xlsx_pivot_unmerge_sheet.delete_cols(3)

list = []
empty_cnt= 0
for i in range(1, item_query_T_xlsx_pivot_unmerge_sheet.max_column+1):
    if(item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=i).value is None):
        cnt = [[i,empty_cnt]]
        list.extend(cnt)
        empty_cnt = empty_cnt+1

for i in range(0, len(list)):
    item_query_T_xlsx_pivot_unmerge_sheet.delete_cols(list[i][0]-list[i][1])


item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=8).value = 'A0202'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=9).value = 'A0302'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=10).value = 'A0402'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=11).value = 'A0502'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=12).value = 'A0602'

list = []
empty_cnt = 0
for i in range(1, item_query_T_xlsx_pivot_unmerge_sheet.max_row + 1):
    if (item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=3).value is None):
        cnt = [[i, empty_cnt]]
        list.extend(cnt)
        empty_cnt = empty_cnt + 1

for i in range(0, len(list)):
    item_query_T_xlsx_pivot_unmerge_sheet.delete_rows(list[i][0] - list[i][1])

item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=13).value = 'ASSY IN'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=14).value = 'ASSY OUT'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=15).value = 'SOLDER'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=16).value = 'W/B'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=17).value = 'MOLD'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=18).value = 'TRIM'
item_query_T_xlsx_pivot_unmerge_sheet.cell(row=1, column=19).value = 'OTHER'

for i in range(2,item_query_T_xlsx_pivot_unmerge_sheet.max_row+1):
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=13).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=3).value
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=14).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=12).value
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=15).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=3).value - item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=8).value
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=16).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=4).value - item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=9).value
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=17).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=5).value - item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=10).value
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=18).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=7).value - item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=12).value
    item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=19).value = item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=6).value - item_query_T_xlsx_pivot_unmerge_sheet.cell(row=i, column=11).value

item_query_T_xlsx_pivot_unmerge.save(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge'+'.xlsx'))

# item_query_T_xlsx_pivot_unmerge A020 기준으로 오름차순 정렬
item_query_T_xlsx_pivot_unmerge_sort(A_source_path,CMP_NM,WW)
item_query_T_xlsx_pivot_unmerge_res_sort(A_source_path,CMP_NM,WW)

# item_query_P_xlsx_pivot_unmerge_res SALE CD 컬럼 삭제를 위해 test_pivot_unmerge_res_sort 엑셀 오픈
item_query_T_xlsx_pivot_unmerge_res_sort = openpyxl.load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge_res_sort'+'.xlsx'))
item_query_T_xlsx_pivot_unmerge_res_sort_sheet = item_query_T_xlsx_pivot_unmerge_res_sort.active

# item_query_P_xlsx_pivot_unmerge_res SALE CD 컬럼 필요없어서 삭제
item_query_T_xlsx_pivot_unmerge_res_sort_sheet.delete_cols(1)
item_query_T_xlsx_pivot_unmerge_res_sort.save(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge_res_sort'+'.xlsx')

# item_query_P_xlsx_pivot_unmerge_res_sort 부분합
item_query_P_xlsx_pivot_unmerge_res_sort_part_sum(A_source_path,CMP_NM,WW)

item_query_T_xlsx_pivot_unmerge_res_sort_part_sum = openpyxl.load_workbook(os.path.abspath(A_source_path+CMP_NM+WW+'주차_test_pivot_unmerge_res_sort_part_sum'+'.xlsx'))
item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum.active

for i in range(2, item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.max_row+1):
    for j in range(2, IXYS_A_Total_sheet.max_row + 1):
        if (item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i,column=1).value == IXYS_A_Total_sheet.cell(row=j,column=4).value):
            IXYS_A_Total_sheet.cell(row=j, column=5).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=2).value
            IXYS_A_Total_sheet.cell(row=j, column=6).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=3).value
            IXYS_A_Total_sheet.cell(row=j, column=8).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=6).value
            IXYS_A_Total_sheet.cell(row=j, column=9).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=8).value
            IXYS_A_Total_sheet.cell(row=j, column=10).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=4).value
            IXYS_A_Total_sheet.cell(row=j, column=11).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=7).value
            IXYS_A_Total_sheet.cell(row=j, column=12).value = item_query_T_xlsx_pivot_unmerge_res_sort_part_sum_sheet.cell(row=i, column=5).value
            IXYS_A_Total_sheet.cell(row=j, column=7).value = str(round((IXYS_A_Total_sheet.cell(row=j, column=6).value/IXYS_A_Total_sheet.cell(row=j,column=5).value)*100,2))+'%'
            IXYS_A_Total_sheet.cell(row=j, column=15).value = str(round((IXYS_A_Total_sheet.cell(row=j, column=14).value / IXYS_A_Total_sheet.cell(row=j, column=13).value) * 100,2)) + '%'

row_for_total = IXYS_A_Total_sheet.max_row
IXYS_A_Total_sheet.merge_cells("A"+str((row_for_total+1))+":D"+str((row_for_total+1)))
IXYS_A_Total_sheet['A'+str(row_for_total+1)] = "TOTAL"
IXYS_A_Total_sheet.cell(row=row_for_total + 1, column=1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')

IXYS_A_Total_sheet.cell(row=row_for_total+1, column=5).value = "=SUM(E2:E"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=6).value = "=SUM(F2:F"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=8).value = "=SUM(H2:H"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=9).value = "=SUM(I2:I"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=10).value = "=SUM(J2:J"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=11).value = "=SUM(K2:K"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=12).value = "=SUM(L2:L"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=13).value = "=SUM(M2:M"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=14).value = "=SUM(N2:N"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=16).value = "=SUM(P2:P"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=17).value = "=SUM(Q2:Q"+str(row_for_total)+")"
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=18).value = "=SUM(R2:R"+str(row_for_total)+")"

ASSY_IN = 0
ASSY_OUT = 0
TEST_IN = 0
TEST_OUT = 0
for i in range(2, row_for_total+1):
    ASSY_IN = ASSY_IN + IXYS_A_Total_sheet.cell(row=i, column=5).value
    ASSY_OUT = ASSY_OUT + IXYS_A_Total_sheet.cell(row=i, column=6).value
    TEST_IN = TEST_IN + IXYS_A_Total_sheet.cell(row=i, column=13).value
    TEST_OUT = TEST_OUT + IXYS_A_Total_sheet.cell(row=i, column=14).value

IXYS_A_Total_sheet.cell(row=row_for_total+1, column=7).value = '=ROUND((F'+str(IXYS_A_Total_sheet.max_row)+'/E'+str(IXYS_A_Total_sheet.max_row)+')*100,2)&"%"'
IXYS_A_Total_sheet.cell(row=row_for_total+1, column=15).value = '=ROUND((N'+str(IXYS_A_Total_sheet.max_row)+'/M'+str(IXYS_A_Total_sheet.max_row)+')*100,2)&"%"'

IXYS_A_Total.save(os.path.abspath(A_source_path+'IXYS test yield repot_WW'+WW+'_2020.xlsx'))
